"""Service de génération de formules de parfum.

Charge les données du coffret (XLSX) une seule fois en mémoire,
puis génère 2 formules personnalisées à partir des réponses utilisateur.
Calcule les quantités en ml pour 3 formats (10ml, 30ml, 50ml).
"""

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from app.data.choice_profile_mapping import (
    CHOICE_PROFILE_MAPPING,
    INGREDIENT_EN_TO_FR,
    PROFILE_DESCRIPTIONS,
    PROFILE_DESCRIPTIONS_EN,
    PROFILE_GENDERS,
)
from app.data.questions import EN_TO_FR_CHOICES
from app.config import get_settings
from app.services import mail_service, redis_service

# ── Chemins ───────────────────────────────────────────────────────────
_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
_XLSX_PATH = _DATA_DIR / "Coffret-description.xlsx"

# ── Cache mémoire (chargé une seule fois) ─────────────────────────────
_coffret: dict | None = None


# ── Normalisation des noms de profils ─────────────────────────────────
# Le XLSX utilise des variantes (Stratégist, Disrupteur, Trail blazer…)
_PROFILE_NORMALIZE = {
    "stratégist": "Strategist",
    "strategist": "Strategist",
    "disrupteur": "Disruptor",
    "disruptor": "Disruptor",
    "trail blazer": "Trailblazer",
    "trailblazer": "Trailblazer",
    "visionary": "Visionary",
    "visionnary": "Visionary",
    "innovator": "Innovator",
    "creator": "Creator",
    "influencer": "Influencer",
    "icon": "Icon",
    "cosy": "Cosy",
}


def _normalize_profile(raw: str | None) -> str | None:
    if not raw:
        return None
    return _PROFILE_NORMALIZE.get(raw.strip().lower(), raw.strip())


# ── Chargement du XLSX ────────────────────────────────────────────────

def _load_coffret() -> dict:
    """Parse le XLSX et retourne les ingrédients + allergènes."""
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)

    # --- Sheet 1 : ingrédients ---
    ws = wb[wb.sheetnames[0]]
    ingredients = []

    # Lignes des ingrédients : T=7-16, C=21-30, F=35-44
    note_ranges = [
        ("top", 7, 16),
        ("heart", 21, 30),
        ("base", 35, 44),
    ]

    for note_type, start, end in note_ranges:
        for row in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=8):
            position = row[0].value
            if not position:
                continue
            ingredients.append({
                "position": position,
                "name": row[1].value,
                "family": row[2].value,
                "description": row[3].value,
                "note_type": note_type,
                "profile_1": _normalize_profile(row[6].value),
                "profile_2": _normalize_profile(row[7].value),
            })

    # --- Sheet 2 : allergènes ---
    ws2 = wb["ALLERGENS"]

    # 3 blocs d'allergènes : top (row 4-31), heart (row 33-60), base (row 62-89)
    allergen_blocks = [
        (4, 6, 31),   # header_row, data_start, data_end
        (33, 35, 60),
        (62, 64, 89),
    ]

    # allergen_map: ingredient_name → set of allergen names
    allergen_map: dict[str, set[str]] = defaultdict(set)

    for header_row, data_start, data_end in allergen_blocks:
        # Read header: col B onwards = ingredient names
        header_cells = list(ws2.iter_rows(
            min_row=header_row, max_row=header_row,
            min_col=2, max_col=11,
        ))[0]
        col_to_ingredient = {}
        for cell in header_cells:
            if cell.value:
                col_to_ingredient[cell.column] = cell.value.strip()

        # Read allergen rows
        for row in ws2.iter_rows(
            min_row=data_start, max_row=data_end,
            min_col=1, max_col=11,
        ):
            allergen_name = row[0].value
            if not allergen_name:
                continue
            allergen_name = allergen_name.strip()
            for cell in row[1:]:
                if cell.value and str(cell.value).strip().lower() == "x":
                    ingredient_name = col_to_ingredient.get(cell.column)
                    if ingredient_name:
                        allergen_map[ingredient_name].add(allergen_name)

    wb.close()

    return {
        "ingredients": ingredients,
        "allergen_map": dict(allergen_map),
    }


def _get_coffret() -> dict:
    global _coffret
    if _coffret is None:
        _coffret = _load_coffret()
    return _coffret


# ── Scoring des profils ──────────────────────────────────────────────

def _score_profiles(answers: dict, gender: str | None) -> list[tuple[str, float]]:
    """Calcule le score de chaque profil à partir des réponses.

    - top_2 choice → +2 pts aux profils associés
    - bottom_2 choice → -1 pt aux profils associés
    - Bonus genre : +1 si le profil match le genre, +0.5 pour unisex
    """
    scores: dict[str, float] = defaultdict(float)

    for qid_str, answer_data in answers.items():
        qid = int(qid_str)
        mapping = CHOICE_PROFILE_MAPPING.get(qid, {})

        if isinstance(answer_data, str):
            answer_data = json.loads(answer_data)

        en_to_fr = EN_TO_FR_CHOICES.get(qid, {})

        def resolve_fr_choice(choice: str) -> str:
            """Exact match first, then prefix match (before ' - ') for descriptive labels."""
            if choice in en_to_fr:
                return en_to_fr[choice]
            choice_lower = choice.lower()
            for key, fr_val in en_to_fr.items():
                key_prefix = key.split(" - ")[0].lower()
                if choice_lower == key_prefix or choice_lower == key.lower():
                    return fr_val
            return choice

        for choice in answer_data.get("top_2", []):
            fr_choice = resolve_fr_choice(choice)
            profiles = mapping.get(fr_choice, [])
            for p in profiles:
                scores[p] += 2.0

        for choice in answer_data.get("bottom_2", []):
            fr_choice = resolve_fr_choice(choice)
            profiles = mapping.get(fr_choice, [])
            for p in profiles:
                scores[p] -= 1.0

    # Bonus genre
    if gender:
        gender_lower = gender.lower()
        # Map "masculin"/"féminin" (FR) to "masculine"/"feminine"
        if gender_lower in ("masculin", "masculine"):
            target = "masculine"
        elif gender_lower in ("féminin", "feminine", "feminin"):
            target = "feminine"
        else:
            target = None

        if target:
            for profile, profile_gender in PROFILE_GENDERS.items():
                if profile_gender == target:
                    scores[profile] += 1.0
                elif profile_gender == "unisex":
                    scores[profile] += 0.5

    # Trier par score décroissant
    return sorted(scores.items(), key=lambda x: x[1], reverse=True)


# ── Sélection des ingrédients ─────────────────────────────────────────

def _get_ingredients_for_profile(
    profile_name: str,
    user_allergens: list[str] | None = None,
) -> dict[str, list[dict]]:
    """Retourne les ingrédients du coffret associés à un profil.

    Filtre les allergènes si l'utilisateur en a déclaré.
    Profile 1 match → priorité haute, Profile 2 → priorité basse.
    """
    coffret = _get_coffret()
    allergen_map = coffret["allergen_map"]

    # Normaliser les allergènes utilisateur pour comparaison insensible à la casse
    blocked_ingredients: set[str] = set()
    if user_allergens:
        user_allergens_lower = {a.strip().lower() for a in user_allergens}
        for ingredient_name, allergens in allergen_map.items():
            allergens_lower = {a.lower() for a in allergens}
            if allergens_lower & user_allergens_lower:
                blocked_ingredients.add(ingredient_name)

    result: dict[str, list[dict]] = {
        "top_notes": [],
        "heart_notes": [],
        "base_notes": [],
    }

    note_type_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}

    for ingredient in coffret["ingredients"]:
        if ingredient["name"] in blocked_ingredients:
            continue

        is_p1 = ingredient["profile_1"] == profile_name
        is_p2 = ingredient["profile_2"] == profile_name

        if is_p1 or is_p2:
            key = note_type_key[ingredient["note_type"]]
            result[key].append({
                "position": ingredient["position"],
                "name": ingredient["name"],
                "family": ingredient["family"],
                "description": ingredient["description"],
                "priority": "primary" if is_p1 else "secondary",
            })

    # Trier : primary d'abord
    for key in result:
        result[key].sort(key=lambda x: 0 if x["priority"] == "primary" else 1)

    return result


# ── Boosters ─────────────────────────────────────────────────────────

BOOSTERS = [
    {
        "name": "Floral",
        "keywords": ["fleur", "rose", "jasmin", "muguet", "floral", "flower",
                      "pétale", "bouquet", "pivoine", "iris", "ylang",
                      "néroli", "magnolia", "tubéreuse", "gardénia"],
    },
    {
        "name": "Ambre doux",
        "keywords": ["ambre", "vanille", "oriental", "chaud", "doux", "warm",
                      "amber", "gourmand", "caramel", "miel", "tonka",
                      "baume", "résine", "encens", "oud", "boisé"],
    },
    {
        "name": "Musc blanc sec",
        "keywords": ["musc", "propre", "frais", "clean", "musk", "coton",
                      "savon", "linge", "poudré", "aldéhyde", "blanc",
                      "minéral", "ozonic", "aquatique", "agrume",
                      "bergamote", "citron", "pamplemousse"],
    },
]


def _select_boosters(
    ingredients: dict[str, list[dict]],
    count: int = 2,
) -> list[dict]:
    """Sélectionne les meilleurs boosters par scoring de mots-clés.

    Analyse les noms, familles et descriptions de toutes les notes du profil,
    puis retourne les `count` boosters avec les meilleurs scores.
    """
    # Construire un texte combiné de toutes les notes pour le matching
    text_parts: list[str] = []
    for note_key in ("top_notes", "heart_notes", "base_notes"):
        for note in ingredients.get(note_key, []):
            for field in ("name", "family", "description"):
                val = note.get(field)
                if val:
                    text_parts.append(val.lower())
    combined = " ".join(text_parts)

    scored: list[tuple[dict, int]] = []
    for booster in BOOSTERS:
        score = sum(1 for kw in booster["keywords"] if kw in combined)
        scored.append((booster, score))

    scored.sort(key=lambda x: x[1], reverse=True)
    return [b for b, _ in scored[:count]]


# ── Allocation des quantités en ml ───────────────────────────────────

# Sélection des notes : basée sur le format 30ml (référence unique)
_NOTE_SELECTION = {
    "top":   {"max_notes": 3},
    "heart": {"max_notes": 3},
    "base":  {"max_notes": 2},
}
_BOOSTER_COUNT = 2  # nombre de boosters sélectionnés

# Configuration des ml par taille (les MÊMES notes pour chaque taille)
_SIZE_CONFIGS = {
    10: {
        "top_ml":     1.0,
        "heart_ml":   1.0,
        "base_ml":    2.0,
        "booster_ml": 1.0,
    },
    30: {
        "top_ml":     4.0,
        "heart_ml":   2.0,
        "base_ml":    1.0,
        "booster_ml": 5.0,
    },
    50: {
        "top_ml":     7.0,
        "heart_ml":   3.0,
        "base_ml":    2.0,
        "booster_ml": 8.0,
    },
}

# Priorités de répartition du déficit par taille
_DEFICIT_PRIORITIES = {
    30: {"top": 0.50, "heart": 0.30, "base": 0.20},
    10: {"top": 0.10, "heart": 0.30, "base": 0.60},
    50: {"top": 0.50, "heart": 0.30, "base": 0.20},
}


def _select_notes(
    ingredients: dict[str, list[dict]],
    boosters: list[dict],
) -> tuple[dict[str, list[dict]], list[dict]]:
    """Sélectionne les notes et boosters une seule fois (identiques pour les 3 tailles)."""
    selected_notes: dict[str, list[dict]] = {}
    for note_type, key in [("top", "top_notes"), ("heart", "heart_notes"), ("base", "base_notes")]:
        max_n = _NOTE_SELECTION[note_type]["max_notes"]
        selected_notes[key] = ingredients.get(key, [])[:max_n]

    selected_boosters = boosters[:_BOOSTER_COUNT]
    return selected_notes, selected_boosters


def _compute_quantities(
    selected_notes: dict[str, list[dict]],
    selected_boosters: list[dict],
    target_ml: int,
) -> dict:
    """Calcule les ml pour chaque note et booster afin d'atteindre le volume cible.

    Les notes sont les MÊMES pour toutes les tailles — seuls les ml changent.
    """
    config = _SIZE_CONFIGS[target_ml]
    priorities = _DEFICIT_PRIORITIES[target_ml]

    ml_map = {
        "top_notes": config["top_ml"],
        "heart_notes": config["heart_ml"],
        "base_notes": config["base_ml"],
    }

    # Attribuer les ml initiaux à chaque note (sans tronquer)
    note_quantities: dict[str, list[dict]] = {}
    for key in ("top_notes", "heart_notes", "base_notes"):
        note_quantities[key] = [
            {**note, "ml": ml_map[key]}
            for note in selected_notes.get(key, [])
        ]

    # Attribuer les ml aux boosters
    booster_quantities = [
        {"name": b["name"], "ml": config["booster_ml"]}
        for b in selected_boosters
    ]

    # Calculer le total actuel
    notes_total = sum(
        n["ml"] for key in ("top_notes", "heart_notes", "base_notes")
        for n in note_quantities[key]
    )
    boosters_total = sum(b["ml"] for b in booster_quantities)
    current_total = notes_total + boosters_total

    # Ajuster pour atteindre le volume cible
    deficit = target_ml - current_total

    if abs(deficit) > 0.01:
        if deficit > 0:
            # Il manque du volume → répartir le déficit selon les priorités
            for note_type, key in [("top", "top_notes"), ("heart", "heart_notes"), ("base", "base_notes")]:
                share = deficit * priorities[note_type]
                notes_in_category = note_quantities[key]
                if notes_in_category:
                    per_note = share / len(notes_in_category)
                    for note in notes_in_category:
                        note["ml"] = round(note["ml"] + per_note, 1)
        else:
            # Excédent → réduction proportionnelle uniforme
            ratio = target_ml / current_total
            for key in ("top_notes", "heart_notes", "base_notes"):
                for note in note_quantities[key]:
                    note["ml"] = round(note["ml"] * ratio, 1)
            for b in booster_quantities:
                b["ml"] = round(b["ml"] * ratio, 1)

    # Compensation finale : corriger l'arrondi sur le premier booster
    final_total = sum(
        n["ml"] for key in ("top_notes", "heart_notes", "base_notes")
        for n in note_quantities[key]
    ) + sum(b["ml"] for b in booster_quantities)

    rounding_error = round(target_ml - final_total, 1)
    if abs(rounding_error) > 0.01 and booster_quantities:
        booster_quantities[0]["ml"] = round(
            booster_quantities[0]["ml"] + rounding_error, 1
        )

    return {
        "target_ml": target_ml,
        "top_notes": note_quantities["top_notes"],
        "heart_notes": note_quantities["heart_notes"],
        "base_notes": note_quantities["base_notes"],
        "boosters": booster_quantities,
    }


# ── Génération des formules ───────────────────────────────────────────

def generate_formulas(session_id: str) -> dict:
    """Génère 2 formules personnalisées pour une session.

    1. Récupère réponses + profil depuis Redis
    2. Score les profils
    3. Sélectionne les ingrédients pour les 2 meilleurs profils
    4. Retourne les formules
    """
    # Récupérer les données
    session_data = redis_service.get_session_answers(session_id)
    if not session_data or not session_data.get("answers"):
        return {"error": "Aucune réponse trouvée", "formulas": []}

    # Déterminer la langue de la session
    session_meta = redis_service.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    descriptions = PROFILE_DESCRIPTIONS_EN if language == "en" else PROFILE_DESCRIPTIONS
    translate_name = (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr" else (lambda name: name)

    profile = redis_service.get_user_profile(session_id)
    gender = profile.get("gender") if profile else None
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    # Parser les allergènes
    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [a.strip() for a in user_allergens_raw.replace(",", ";").split(";") if a.strip()]

    # Scorer les profils
    ranked = _score_profiles(session_data["answers"], gender)

    if len(ranked) < 2:
        return {"error": "Not enough data to generate formulas" if language == "en" else "Pas assez de données pour générer des formules", "formulas": []}

    # Prendre les 2 meilleurs, en s'assurant qu'ils ont des ingrédients
    top_profiles = []
    for profile_name, score in ranked:
        if len(top_profiles) >= 2:
            break
        ingredients = _get_ingredients_for_profile(profile_name, user_allergens)
        has_notes = any(ingredients[k] for k in ("top_notes", "heart_notes", "base_notes"))
        if not has_notes:
            continue
        top_profiles.append((profile_name, score))

    if len(top_profiles) < 2:
        return {"error": "Not enough profiles with ingredients to generate formulas" if language == "en" else "Pas assez de profils avec des ingrédients pour générer des formules", "formulas": []}

    # Générer les formules
    formulas = []
    for profile_name, score in top_profiles:
        ingredients = _get_ingredients_for_profile(profile_name, user_allergens)

        # Sélectionner les boosters pour ce profil
        boosters = _select_boosters(ingredients)

        # Traduire les noms d'ingrédients dans les détails
        translated_details = {}
        for note_key in ("top_notes", "heart_notes", "base_notes"):
            translated_details[note_key] = [
                {**n, "name": translate_name(n["name"])} for n in ingredients[note_key]
            ]

        # Sélectionner les notes une seule fois (identiques pour les 3 tailles)
        selected_notes, selected_boosters = _select_notes(translated_details, boosters)

        # Calculer les quantités pour les 3 formats (mêmes notes, ml différents)
        sizes = {}
        for target_ml in (10, 30, 50):
            sizes[f"{target_ml}ml"] = _compute_quantities(
                selected_notes, selected_boosters, target_ml
            )

        formulas.append({
            "profile": profile_name,
            "description": descriptions.get(profile_name, ""),
            "score": score,
            "top_notes": [n["name"] for n in translated_details["top_notes"]],
            "heart_notes": [n["name"] for n in translated_details["heart_notes"]],
            "base_notes": [n["name"] for n in translated_details["base_notes"]],
            "details": translated_details,
            "sizes": sizes,
        })

    # Stocker les formules générées dans Redis pour pouvoir les sélectionner ensuite
    redis_service.save_generated_formulas(session_id, formulas)

    return {"formulas": formulas}


# ── Sélection et personnalisation ─────────────────────────────────────

def select_formula(session_id: str, formula_index: int) -> dict:
    """Sélectionne une des 2 formules générées et la stocke dans Redis."""
    formulas = redis_service.get_generated_formulas(session_id)
    if not formulas:
        return {"error": "No generated formulas found"}
    if formula_index not in (0, 1):
        return {"error": "formula_index must be 0 or 1"}
    if formula_index >= len(formulas):
        return {"error": "Invalid formula index"}

    selected = formulas[formula_index]
    redis_service.save_selected_formula(session_id, selected)

    internal_email = get_settings().internal_email
    if internal_email:
        try:
            mail_service.send_mail(internal_email, session_id, selected)
        except Exception as e:
            print(f"[mail] Erreur lors de l'envoi interne : {e}")

    return {"formula": selected}


def get_available_ingredients(
    session_id: str,
    note_type: str,
) -> dict:
    """Retourne tous les ingrédients disponibles pour un type de note, filtré par allergènes."""
    if note_type not in ("top", "heart", "base"):
        return {"error": "note_type must be top, heart, or base"}

    profile = redis_service.get_user_profile(session_id)
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [a.strip() for a in user_allergens_raw.replace(",", ";").split(";") if a.strip()]

    # Déterminer la langue pour la traduction
    session_meta = redis_service.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    translate_name = (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr" else (lambda name: name)

    coffret = _get_coffret()
    allergen_map = coffret["allergen_map"]

    blocked_ingredients: set[str] = set()
    if user_allergens:
        user_allergens_lower = {a.strip().lower() for a in user_allergens}
        for ingredient_name, allergens in allergen_map.items():
            allergens_lower = {a.lower() for a in allergens}
            if allergens_lower & user_allergens_lower:
                blocked_ingredients.add(ingredient_name)

    note_type_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}
    key = note_type_key[note_type]

    ingredients = []
    for ingredient in coffret["ingredients"]:
        if ingredient["note_type"] != note_type:
            continue
        if ingredient["name"] in blocked_ingredients:
            continue
        ingredients.append({
            "name": translate_name(ingredient["name"]),
            "family": ingredient["family"],
            "description": ingredient["description"],
        })

    return {"note_type": note_type, "ingredients": ingredients}


def replace_note(
    session_id: str,
    note_type: str,
    old_note: str,
    new_note: str,
) -> dict:
    """Remplace une note dans la formule sélectionnée et recalcule les ml."""
    if note_type not in ("top", "heart", "base"):
        return {"error": "note_type must be top, heart, or base"}

    selected = redis_service.get_selected_formula(session_id)
    if not selected:
        return {"error": "No formula selected yet"}

    note_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}[note_type]

    # Trouver la nouvelle note dans le coffret
    session_meta = redis_service.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    translate_name = (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr" else (lambda name: name)

    coffret = _get_coffret()
    new_ingredient = None
    for ingredient in coffret["ingredients"]:
        if ingredient["note_type"] != note_type:
            continue
        translated = translate_name(ingredient["name"])
        if translated.lower() == new_note.lower() or ingredient["name"].lower() == new_note.lower():
            new_ingredient = {
                "name": translated,
                "family": ingredient["family"],
                "description": ingredient["description"],
                "position": ingredient["position"],
                "priority": "custom",
            }
            break

    if not new_ingredient:
        return {"error": f"Ingredient '{new_note}' not found in coffret for {note_type} notes"}

    # Remplacer dans les détails
    details = selected.get("details", {})
    found = False
    for i, note in enumerate(details.get(note_key, [])):
        if note["name"].lower() == old_note.lower():
            details[note_key][i] = new_ingredient
            found = True
            break

    if not found:
        return {"error": f"Note '{old_note}' not found in current formula's {note_key}"}

    # Mettre à jour la liste simplifiée
    selected[note_key] = [n["name"] for n in details[note_key]]
    selected["details"] = details

    # Recalculer les ml pour les 3 tailles (mêmes notes partout)
    boosters_data = selected.get("sizes", {}).get("30ml", {}).get("boosters", [])
    boosters = [{"name": b["name"], "keywords": []} for b in boosters_data]

    selected_notes, selected_boosters = _select_notes(details, boosters)

    sizes = {}
    for target_ml in (10, 30, 50):
        sizes[f"{target_ml}ml"] = _compute_quantities(selected_notes, selected_boosters, target_ml)

    selected["sizes"] = sizes

    # Sauvegarder dans Redis
    redis_service.save_selected_formula(session_id, selected)

    return {"formula": selected}
